import os
import pandas as pd
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter


class Export():
    def __init__(self) -> None:
        self.access_policy_header = ['Name', 'Count Of Enabled Rules', 'Count Of Allowed Rules', 'Ratio Of Enabled Rules', 'Ratio Of Allowed Rules', 'Average Source Network Size', 'Average Source Network Size /', 'Average Destination Network Size', 'Average Destination Network Size /', 'Average Destination Port Size']
        self.access_rule_header = ['Name', 'Action', 'Enabled', 'Source Zones', 'Source Networks', 'Source Ports', 'Destination Zones', 'Destination Networks', 'Destination Ports', 'Source Networks Size', 'Source Networks Size /', 'Destination Networks Size', 'Destination Networks Size /', 'Destination Ports Size', 'Source Network Category', 'Relative Source Network Category', 'Destination Network Category', 'Relative Destination Network Category', 'Destination Port Category', 'Relative Destination Port Category', 'Duplicated', 'Reversed', 'Merge Candidates']
        self.port_header = ['Group Name', 'Name', 'Protocol', 'Port', 'Size', 'Risky', 'Duplicates', 'Reference Count from Rules']
        self.network_header = ['Group Name', 'Group depth', 'Name', 'Value', 'Size', 'Size /', 'Duplicates', 'Reference Count from Rules']

    def export_to_excel(self, data: list[str], header: list[str], dir: str, file_name: str, sheet_name: str) -> None:
        """Export the specified data to excel.

            Args:
            ----
                data: The specified data to export.
                header: The specified header for data
                sheet_name: The name of the sheet

        """
        df = pd.DataFrame(data, columns=header)
        df.index = range(1, len(df)+1)
        export_dir = os.getcwd() + '/' + dir
        if not os.path.exists(export_dir):
            os.mkdir(export_dir)
        export_path = export_dir + '/' + file_name
        if os.path.exists(export_path):
            with pd.ExcelWriter(path=export_path, mode='a', if_sheet_exists='replace', engine='openpyxl') as writer:
                df.to_excel(excel_writer=writer, sheet_name=sheet_name)
        else:
            with pd.ExcelWriter(path=export_path) as writer:
                df.to_excel(excel_writer=writer, sheet_name=sheet_name)

        wb = openpyxl.load_workbook('./exports/final.xlsx')
        self.__format_column(wb[sheet_name])
        self.__format_row(wb[sheet_name])
        wb.save('./exports/final.xlsx')

    def __format_column(self, ws: Worksheet):
        for i, column in enumerate(ws.iter_cols(), start=1):
            max_length = 0
            for cell in column:
                if '\n' not in str(cell.value):
                    length = len(str(cell.value))
                    if length > max_length:
                        max_length = length
            ws.column_dimensions[get_column_letter(i)].width = (max_length + 2) * 2

    def __format_row(self, ws: Worksheet):
        for i, row in enumerate(ws.iter_rows(), start=1):
            max_line = 0
            for cell in row:
                lines = str(cell.value).count('\n') + 1
                if lines > max_line:
                    max_line = lines
                if i == 1:
                    cell.font = Font(size=16, color="1F497D", bold=True)
                else:
                    cell.font = Font(size=14)
                cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')
            ws.row_dimensions[i].height = 20 * max_line